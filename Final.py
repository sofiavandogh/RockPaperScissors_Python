from tkinter import *
from tkinter.messagebox import *
import random
import winsound
import openpyxl
from openpyxl import Workbook


###############################################################
##GAME LOGIC AND RULES#######################################

#VARIABLES
choices_list = ["Pedra", "Papel", "Tesoura"]
user = 0
computer = 0

wb = openpyxl.load_workbook('winnerecord.xlsx')
sheet = wb.active

#FUNCTIONS###################################################

##Players Options
def chose_rock():
    computer_choice =  "".join(random.sample(choices_list,1))
    
    if computer_choice == "Papel":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que embrulha a Pedra. \nOhhh, perdeste esta jogada!")
        global computer
        computer = computer + 1
        label_computer.config(text = f"Computador: {computer}")
    elif computer_choice == "Tesoura":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que é esmagada pela Pedra. Wow, ganhaste esta jogada!")
        global user
        user = user + 1
        label_user.config(text = f"Utilizador: {user}")
    else:
        showinfo(title = "Escolha do Computador",
                 message = f"O computador também escolheu {computer_choice}. Ups, é um empate!")
    get_winner()
    
def chose_paper():
    computer_choice =  "".join(random.sample(choices_list,1))
    
    if computer_choice == "Tesoura":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que corta o Papel. \nOhhh, perdeste esta jogada!")
        global computer
        computer = computer + 1
        label_computer.config(text=f"Computador: {computer}")
    elif computer_choice == "Pedra":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que é embrulhada pelo Papel. \nWow, ganhaste esta jogada!")
        global user
        user = user + 1
        label_user.config(text = f"Utilizador: {user}")
    else:
        showinfo(title = "Escolha do Computador",
                 message = f"O computador também escolheu {computer_choice}. Ups, é um empate!")
    get_winner()
    
def chose_scissor():
    computer_choice = "".join(random.sample(choices_list,1))
    
    if computer_choice == "Pedra":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que esmaga a Tesoura. \nOhhh, perdeste esta jogada!")
        global computer
        computer = computer + 1
        label_computer.config(text = f"Computador: {computer}")
    elif computer_choice == "Papel":
        showinfo(title = "Escolha do Computador",
                 message = f"O computador escolheu {computer_choice}, que é cortado pela Tesoura. \nWow, ganhaste esta jogada!")
        global user
        user = user + 1
        label_user.config(text = f"Utilizador: {user}")
    else:
        showinfo(title = "Escolha do Computador",
                 message = f"O computador também escolheu {computer_choice}. Ups, é um empate!")
    get_winner()

#The first to reach 3 points wins Game
def get_winner():
    if user == 3 and computer < 3:
        label_user_winner.pack()
        winsound.PlaySound("tada.wav", winsound.SND_ASYNC)
        showinfo(title = "Game Over", message = "Parabéns, ganhaste o Jogo!")
        resetUser()
    elif user <3 and computer == 3:
        label_computer_winner.pack()
        winsound.PlaySound("loser.wav", winsound.SND_ASYNC)
        showinfo(title = "Game Over", message = "Ohhh, que pena... perdeste o Jogo!")
        resetComputer()

##Register Winner in Excel Sheet
def registerUser():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row = current_row + 1, column = 1).value = "Utilizador"
    wb.save('winnerecord.xlsx')

def registerComputer():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row = current_row + 1, column = 1).value = "Computador"
    wb.save('winnerecord.xlsx')
    
#After Game Over, winner is registered in Excel Sheet and user can play again and reset values
def resetUser():
    registerUser()
       
    answer = askyesno(title = "Novo Jogo?", message = "Queres jogar outra vez?")
    if answer == True:
         global user
         global computer
         user = 0
         computer = 0
         label_user.config(text = f"Utilizador: {user}")
         label_computer.config(text = f"Computador: {user}")
         label_winner.config(text = "E o vencedor é...")
         label_user_winner.pack_forget()
         label_computer_winner.pack_forget()
    else:
         window.destroy()

def resetComputer():
    registerComputer()
       
    answer = askyesno(title = "Novo Jogo?", message = "Queres jogar outra vez?")
    if answer == True:
         global user
         global computer
         user = 0
         computer = 0
         label_user.config(text = f"Utilizador: {user}")
         label_computer.config(text = f"Computador: {user}")
         label_winner.config(text = "E o vencedor é...")
         label_user_winner.pack_forget()
         label_computer_winner.pack_forget()
    else:
         window.destroy()
         
#######################################################
##LAYOUT AND DESIGN##################################
         
##WINDOW
window = Tk()
window.title("Python II Final Project")
window.resizable(False, False)
window["bg"] = "#325ca8"
window.iconbitmap("mylogo.ico")

win_w = 1000 
win_h = 750  

screen_w = window.winfo_screenwidth() 
screen_h = window.winfo_screenheight() 

pos_x = (screen_w / 2) - (win_w / 2)  
pos_y = (screen_h / 2) - (win_h / 2) 

window.geometry("%dx%d+%d+%d" % (win_w,win_h,pos_x,pos_y))

##WIDGETS/OBJECTS#####################################
##FRAMES
frame_btn = Frame(window, background="#325ca8")
frame_score = Frame(frame_btn, background="white")

##LABELS
label_title = Label(window, text = "O Grande Jogo", bg = "#325ca8",
                    fg = "white")
label_subtitle = Label(window, text = "PEDRA, PAPEL E TESOURA", bg = "white",
                    fg = "#325ca8")
label_play = Label(window, text = "Escolha a sua próxima jogada: ", bg = "#325ca8", fg = "white")
label_winner = Label(window, text = "E o vencedor é... ", bg = "#325ca8",
                     fg = "white")
label_rule = Label(window, text = "(O primeiro a alcançar 3 pontos)", bg = "#325ca8",
                     fg = "white")
label_score = Label(frame_score, text = "PONTUAÇÃO", bg = "white",
                    fg = "#000000")
label_user = Label(frame_score, text = "Utilizador: 0", bg = "white",
                   fg = "#325ca8")
label_computer = Label(frame_score, text = "Computador: 0", bg = "white",
                    fg = "#3e71cf")
user_winner_image = PhotoImage(file = "user.png")
label_user_winner = Label(window, image = user_winner_image, text = "Utilizador",
                    compound = BOTTOM, bg = "#325ca8", fg = "black")

computer_winner_image = PhotoImage(file = "computer.png")
label_computer_winner = Label(window, image = computer_winner_image, text = "Computador",
                    compound = BOTTOM, bg = "#325ca8", fg = "black")

font_title = ("Calibri", 40, "bold", "italic")
font_subtitle = ("Calibri", 50, "bold")
font_play = ("Calibri", 20, "bold")
font_rule = ("Calibri", 10, "italic")
font_btn = ("Calibri", 15, "bold")

label_title.configure(font = font_title)
label_subtitle.configure(font = font_subtitle)
label_play.configure(font = font_play)
label_winner.configure(font = font_play)
label_rule.configure(font = font_rule)
label_score.configure(font = font_btn)
label_user.configure(font = font_btn)
label_computer.configure(font = font_btn)
label_user_winner.configure(font = font_btn)
label_computer_winner.configure(font = font_btn)

##BUTTONS
rock_image = PhotoImage(file = "rock.png")
btn_rock = Button(frame_btn, text = "PEDRA", image = rock_image, command = chose_rock,
                  compound = TOP,  bg = "#325ca8", fg = "white", font = font_btn,  bd = 5)
 
paper_image = PhotoImage(file = "paper.png")
btn_paper = Button(frame_btn, text = "PAPEL", image = paper_image, command = chose_paper,
                   compound = TOP, bg = "#325ca8", fg = "white", font = font_btn, bd = 5)

scissor_image = PhotoImage(file = "scissor.png")
btn_scissor = Button(frame_btn, text = "TESOURA", image = scissor_image, command = chose_scissor,
                     compound = TOP, bg = "#325ca8", fg = "white", font = font_btn, bd = 5)

##LAYOUT POSITIONING###################################
label_title.pack(fill = X, ipady = 15)
label_subtitle.pack(fill = X, ipady = 10, pady = 20)
label_play.pack(ipady = 15)
label_score.pack()
label_user.pack()
label_computer.pack()

btn_rock.pack(side = LEFT, ipadx = 10, ipady = 5)
btn_paper.pack(side = LEFT, padx = 30,  ipadx = 10, ipady = 5)
btn_scissor.pack(side = LEFT,  ipadx = 10, ipady = 5)

frame_score.pack(side = LEFT, padx = 20, pady = 20, ipadx = 20, ipady = 5)
frame_btn.pack(ipadx = 10, ipady = 20)

label_winner.pack()
label_rule.pack()

window.mainloop()


